from rest_framework.viewsets import GenericViewSet
from rest_framework import mixins
import traceback
from general.models import UserProfile
from rest_framework import status
from rest_framework.mixins import ListModelMixin, RetrieveModelMixin
from rest_framework.parsers import MultiPartParser, FormParser
from english.models import Student, Test, TestQuestion, Result
from english.serializers import StudentSerializer, TestSerializer, TestQuestionSerializer, ResultSerializer
from rest_framework.decorators import action
from rest_framework.response import Response
from english.models import Student, Test, TestQuestion, TestQuestionVariant, Result, Tutor
from .serializers import StatsSerializer, StudentSerializer, TestQuestionSerializer, TestSerializer, ResultSerializer, TestQuestionVariantSerializer, TutorSerializer
from django.contrib.auth.models import User
from english.permissions import(
    CanSeeAllStudentsPermission, CanCreateStudentsPermission, CanUpdateStudentsPermission, CanDeleteStudentsPermission, CanSeeHimselfPermission,
    CanSeeTestQuestionsPermission, CanCreateTestQuestionsPermission, CanDeleteTestQuestionsPermission, CanUpdateTestQuestionsPermission,
    CanSeeTestQuestionsVariantsPermission, CanCreateTestQuestionsVariantsPermission, CanDeleteTestQuestionsVariantsPermission, CanUpdateTestQuestionsVariantsPermission,
    CanSeeAllResultsPermission, CanSeeOwnResultPermission,
    CanSeeAllTutorsPermission, CanSeeHisTutorPermission, CanCreateTutorsPermission, CanDeleteTutorsPermission, CanUpdateTutorsPermission, SecondFactorPermission)
from rest_framework import permissions
import io
from datetime import datetime
from django.http import HttpResponse
from django.db.models import Avg, Count, Max,Min   
import openpyxl 


class StudentsViewSet(
    mixins.CreateModelMixin, 
    mixins.ListModelMixin, 
    mixins.RetrieveModelMixin, 
    mixins.DestroyModelMixin,
    mixins.UpdateModelMixin,
    GenericViewSet
):
    @action(detail=False, methods=["GET"], url_path="export")
    def export(self, request, *args, **kwargs):
            bearings = self.get_queryset()
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Студенты"
            headers = [
                "ID", 
                "Имя", 
                "Прогресс", 
            ]
            
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
            
            for row_num, bearing in enumerate(bearings, 2):
                ws.cell(row=row_num, column=1, value=bearing.id)
                ws.cell(row=row_num, column=2, value=bearing.name)
                ws.cell(row=row_num, column=3, value=bearing.progress)
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            filename = f"bearings_export_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
            
            response = HttpResponse(
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
        

    
    queryset = Student.objects.all()
    serializer_class = StudentSerializer    
    #permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()
        if user.is_superuser:
            return qs
        elif hasattr(user, 'userprofile') and user.userprofile.type == 'tutor':
            return qs.filter(tutor=user)
        else:
            return qs.none()
        
    def get_permissions(self):
        if self.action == 'create':
            return [CanCreateStudentsPermission(), SecondFactorPermission()]

        elif self.action == 'update':
            return [CanUpdateStudentsPermission(), SecondFactorPermission()]

        elif self.action == 'destroy':
            return [CanDeleteStudentsPermission(), SecondFactorPermission()]
        elif self.action == 'stats': 
            return [SecondFactorPermission()]
        else:
            return [permissions.IsAuthenticated()]
        
    @action(detail=True, methods=['POST'], url_path='assign-tutor')
    def assign_tutor(self, request, pk=None):
        try:
            student = self.get_object()
            if not request.user.is_superuser:
                return Response({"error": "Только суперюзер может назначать тьютора"}, status=403)

            tutor_id = request.data.get('tutor_id')
            if not tutor_id:
                return Response({"error": "tutor_id обязателен"}, status=400)

            tutor_profile = UserProfile.objects.filter(id=tutor_id, type='tutor').first()
            if not tutor_profile:
                return Response({"error": "Преподаватель не найден"}, status=404)

            student.tutor = tutor_profile.user
            student.save()
            return Response({"success": True})
        except Exception as e:
            return Response({"error": str(e), "trace": traceback.format_exc()}, status=500)


    @action(detail=True,methods=['POST'],url_path='assign-tests' )
    def assign_tests(self, request, pk=None):
        student = self.get_object()
        user = request.user

        is_superuser = user.is_superuser
        is_tutor = (hasattr(user, 'userprofile') and user.userprofile.type == 'tutor' and student.tutor == user)

        if not (is_superuser or is_tutor):
            return Response(
                {"error": "Нет прав на назначение тестов"},
                status=status.HTTP_403_FORBIDDEN
            )

        test_ids = request.data.get('assigned_tests', [])

        if not isinstance(test_ids, list):
            return Response(
                {"error": "assigned_tests должен быть списком"},
                status=status.HTTP_400_BAD_REQUEST
            )
        student.assigned_tests.set(test_ids)
        student.save()

        return Response({"success": True})
   # @action(detail=False, url_path="create-custom", methods=['POST'], permission_classes=[CanCreateStudentsPermission])
   # def create_custom_group(self, request, *args, **kwargs):
    #    student = Student.objects.create(progress=0)
     #   return Response({
      #      "success": True
       # })
       
    @action(detail=False, url_path="create-custom-group", methods=['POST'], permission_classes=[SecondFactorPermission])
    def create_custom_group(self, request, *args, **kwargs):
        student = Student.objects.create()
        return Response({
            "success": True
        })
        
    @action(detail=False, url_path="stats", methods=['GET'], permission_classes=[SecondFactorPermission])
    def get_stats(self, request, *args, **kwargs):
     stats = Student.objects.aggregate(count=Count("*"), avg=Avg("id"), max=Max("id"), min=Min("id"))
     return Response(stats)
       
     

class TestsViewSet(mixins.CreateModelMixin,
                   mixins.UpdateModelMixin,
                   mixins.RetrieveModelMixin,
                   mixins.ListModelMixin,
                   GenericViewSet):
    queryset = Test.objects.all()
    serializer_class = TestSerializer
    
    @action(detail=False, url_path="stats", methods=['GET'], permission_classes=[SecondFactorPermission])
    def get_stats(self, request, *args, **kwargs):
     stats = Test.objects.aggregate(count=Count("*"), avg=Avg("id"), max=Max("id"), min=Min("id"))
     return Response(stats)

class TestQuestionsViewSet(mixins.CreateModelMixin,
                           mixins.UpdateModelMixin,
                           mixins.RetrieveModelMixin,
                           mixins.ListModelMixin,
                           mixins.DestroyModelMixin,
                           GenericViewSet):
    queryset = TestQuestion.objects.all()
    serializer_class = TestQuestionSerializer
    
    def get_queryset(self):
        qs = TestQuestion.objects.all()
        test_id = self.request.query_params.get('test')
        if test_id:
            qs = qs.filter(test_id=test_id)
        return qs
    
    @action(detail=False, url_path="stats", methods=['GET'], permission_classes=[SecondFactorPermission])
    def get_stats(self, request, *args, **kwargs):
     stats = TestQuestion.objects.aggregate(count=Count("*"), avg=Avg("id"), max=Max("id"), min=Min("id"))
     return Response(stats)

class ResultsViewSet(mixins.CreateModelMixin,
                     mixins.UpdateModelMixin,
                     mixins.RetrieveModelMixin,
                     mixins.ListModelMixin,
                     GenericViewSet):
    queryset = Result.objects.all()
    serializer_class = ResultSerializer
    
    @action(detail=False, url_path="stats", methods=['GET'], permission_classes=[SecondFactorPermission])
    def get_stats(self, request, *args, **kwargs):
     stats = Result.objects.aggregate(count=Count("*"), avg=Avg("id"), max=Max("id"), min=Min("id"))
     return Response(stats)

class TestQuestionVariantViewSet(
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    mixins.ListModelMixin,
    GenericViewSet
):
    queryset = TestQuestionVariant.objects.all()
    serializer_class = TestQuestionVariantSerializer
    
    @action(detail=False, url_path="stats", methods=['GET'])
    def get_stats(self, request, *args, **kwargs):
     stats = TestQuestionVariant.objects.aggregate(count=Count("*"), avg=Avg("id"), max=Max("id"), min=Min("id"))
     return Response(stats)
